#!/usr/bin/env python3
from __future__ import annotations

import argparse, io, os, re, sys
from calendar import monthrange
from datetime import datetime
from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pdfplumber

import annual_kpi_report as core
from src.auth.client import login
from src.integrations.google_drive import find_file, upload_xlsx_as_google_sheet
from src.integrations.google_sheets import ensure_hidden_sheet, read_values, sheet_id, sheets_service, write_values

SHEET_TITLE="Annual KPI"; META_TITLE="_META"; MANUAL="MANUAL INPUT REQUIRED"
MONTHS=["Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","Jan","Feb","Mar"]
HEADERS=["KPI","Year"]+MONTHS+["","Upto"]
FIXED=["HSD KMPL INCL AC","HSD KMPL EXCL AC","TOTAL LUB KMPL","B.D RATE","MED CANCL.","SPRING CONS","AVG TYRE LIFE","NEW TYRE LIFE","RC TYRE LIFE","N.T.S RATE","Ist RC S Rate","TTL SCP Rate","RT Factor"]
TYRE=FIXED[-7:]
TYRE_SITE={
"BADVEL":("BDV","KADAPA(KDP ZONE)","DPTO YSR KADAPA"),"JAMMALAMADUGU":("JMD","KADAPA(KDP ZONE)","DPTO YSR KADAPA"),"KADAPA":("KDP","KADAPA(KDP ZONE)","DPTO YSR KADAPA"),"MYDUKUR":("MYD","KADAPA(KDP ZONE)","DPTO YSR KADAPA"),"PRODDUTUR":("PDT","KADAPA(KDP ZONE)","DPTO YSR KADAPA"),"PULIVENDULA":("PVD","KADAPA(KDP ZONE)","DPTO YSR KADAPA"),"RAJAMPET":("RJP","KADAPA(KDP ZONE)","DPTO YSR KADAPA")}
REPORTS=Path(__file__).resolve().parent/"reports"; REPORTS.mkdir(exist_ok=True)

def n(v): return core.norm(v)
def num(v): return core.number(v)
def selected_fy(y,m):
    s=y if m>=4 else y-1; return f"{s}-{str(s+1)[-2:]}"
def fy_triplet(y,m):
    s=y if m>=4 else y-1; return [f"{x}-{str(x+1)[-2:]}" for x in (s-2,s-1,s)]
def fy_months(fy,sy,sm):
    s,e=core.parse_fy(fy); out=[(s,m) for m in range(4,13)]+[(e,m) for m in range(1,4)]
    return [(y,m) for y,m in out if fy!=selected_fy(sy,sm) or (y,m)<=(sy,sm)]
def month_name(m): return MONTHS[(m-4)%12]
def token(y,m,sep="_"): return f"{y}{m:02d}{datetime(y,m,1).strftime('%B')}{sep}{y}"
def district(region): return n(region).replace(" ","")
def at(row,i): return num(row[i]) if i is not None and i<len(row) else None

def get_html(s,base,path,params):
    r=s.get(f"{base}/{path}",params=params,timeout=45); r.raise_for_status()
    if "<table" not in r.text.lower(): raise RuntimeError(f"{path} returned no table")
    return r.text
def post_html(s,base,path,data):
    r=s.post(f"{base}/{path}",data=data,timeout=45); r.raise_for_status()
    if "<table" not in r.text.lower(): raise RuntimeError(f"{path} returned no table")
    return r.text

def find_depot(html,display,vehicle,required):
    for table in BeautifulSoup(html,"html.parser").find_all("table"):
        text=n(table.get_text(" ",strip=True))
        if not all(n(x) in text for x in required): continue
        h,rows=core.expanded_headers(table); row=core.depot_row(h,rows,display,vehicle)
        if row: return h,row
    return None,None
def idxs(h,must,reject=()): return [i for i,x in enumerate(h) if all(n(a) in n(x) for a in must) and not any(n(a) in n(x) for a in reject)]

def fetch_hsd(s,d,v,r,y,m):
    last=monthrange(y,m)[1]
    h,row=find_depot(get_html(s,core.MED_BASE,"mth_acnac_dpt.php",{"action":"","fdate":f"{last:02d}/{m:02d}/{y}","dist":district(r)}),d,v,["WITH AC KMPL","WITHOUT AC KMPL"])
    if not row:return {}
    wf=idxs(h,["WITH AC KMPL","FOR"],["WITHOUT"]); wu=idxs(h,["WITH AC KMPL","UP TO"],["WITHOUT"])
    nf=idxs(h,["WITHOUT AC KMPL","FOR"]); nu=idxs(h,["WITHOUT AC KMPL","UP TO"])
    return {"HSD KMPL INCL AC":{"month":at(row,wf[0] if wf else None),"upto":at(row,wu[0] if wu else None)},"HSD KMPL EXCL AC":{"month":at(row,nf[0] if nf else None),"upto":at(row,nu[0] if nu else None)}}

def direct_dimension_rows(html,label,prefix):
    # These APSRTC pages have a stable body layout: SNO, PRODUCT/ENGINE TYPE, BUSES/HELD,
    # FOR MONTH CY, LY, VAR, UPTO CY, LY, VAR. Parse body rows directly to avoid colspan header drift.
    soup=BeautifulSoup(html,"html.parser"); out={}
    for table in soup.find_all("table"):
        text=n(table.get_text(" ",strip=True))
        if n(label) not in text or "FOR THE MONTH" not in text or "UP TO THE MONTH" not in text: continue
        for tr in table.find_all("tr"):
            cells=[c.get_text(" ",strip=True) for c in tr.find_all(["td"],recursive=False)]
            if len(cells)<7: continue
            name=cells[1].strip()
            if not name or n(name) in {"TOTAL","GRAND TOTAL"}: continue
            # Defensive rule: source dimension names must contain letters; numeric totals/held counts are never names.
            if not re.search(r"[A-Za-z]",name): continue
            mv=num(cells[3]); uv=num(cells[6])
            out[f"{prefix}: {name}"]={"month":mv,"upto":uv}
    return out

def fetch_dimension(s,path,label,prefix,d,r,y,m):
    html=post_html(s,core.MEDNEW_BASE,path,{"yymm":token(y,m),"reg":district(r),"dept":d})
    out=direct_dimension_rows(html,label,prefix)
    if not out: raise RuntimeError(f"No valid {prefix.lower()} names parsed")
    return out

def fetch_pair(s,base,path,params,d,v,required,matcher,name):
    h,row=find_depot(get_html(s,base,path,params),d,v,required)
    if not row:return {}
    ii=[i for i,x in enumerate(h) if matcher(n(x))]
    return {name:{"month":at(row,ii[0] if ii else None),"upto":at(row,ii[1] if len(ii)>1 else None)}}
def fetch_bd(s,d,v,r,y,m): return fetch_pair(s,core.MED_BASE,"sysbd_dpt.php",{"action":"","yymm":token(y,m),"dist":district(r)},d,v,["BD RATE"],lambda x:"BD RATE" in x,"B.D RATE")
def fetch_med(s,d,v,r,y,m): return fetch_pair(s,core.MEDNEW_BASE,"medcan_um_dpt.php",{"action":"","fdate":token(y,m,"-"),"dist":district(r)},d,v,["CANC"],lambda x:"%" in x and "CANC" in x,"MED CANCL.")
def fetch_spring(s,d,v,r,y,m): return fetch_pair(s,"http://103.44.14.20/storeap","deptspring.php",{"action":"","yymm":token(y,m),"dist":district(r)},d,v,["SPRING CONSUMPTION PER LAKH KMS"],lambda x:"SPRING CONSUMPTION PER LAKH KMS" in x,"SPRING CONS")

def page_month(html,y,m):
    t=n(BeautifulSoup(html,"html.parser").get_text(" ",strip=True)); mn=datetime(y,m,1).strftime("%B").upper()
    return any(x in t for x in (f"{mn}_{y}",f"{mn}-{y}",f"{mn} {y}"))
def lub_html(s,y,m):
    url=f"{core.MED_BASE}/lub_rgn_rpt.php"; wanted=f"{datetime(y,m,1).strftime('%B')}_{y}"
    landing=s.get(url,timeout=45); landing.raise_for_status(); soup=BeautifulSoup(landing.text,"html.parser")
    # First use the site's own form and exact option value. This is the authoritative POST path.
    for form in soup.find_all("form"):
        data={}
        for inp in form.find_all("input"):
            if inp.get("name") and inp.get("type","").lower() not in {"submit","button"}: data[inp["name"]]=inp.get("value","")
        found=False
        for sel in form.find_all("select"):
            name=sel.get("name")
            if not name: continue
            for opt in sel.find_all("option"):
                ov=opt.get("value",""); ot=opt.get_text(" ",strip=True)
                if n(wanted) in {n(ov),n(ot)}:
                    data[name]=ov or ot; found=True; break
        if found:
            action=form.get("action") or url
            if not action.startswith("http"): action=f"{core.MED_BASE}/{action.lstrip('/')}"
            rr=s.post(action,data=data,timeout=45); rr.raise_for_status()
            if page_month(rr.text,y,m):
                print(f"LUB POST verified {wanted} using fields: {','.join(sorted(data))}")
                return rr.text
    # Do not guess a value. Try common field names only to discover the live form contract, then verify report month.
    for field in ("month_year","yymm","month","mth","fdate","mon"):
        rr=s.post(url,data={field:wanted},timeout=45); rr.raise_for_status()
        if page_month(rr.text,y,m):
            print(f"LUB POST verified {wanted} using field: {field}")
            return rr.text
    raise RuntimeError(f"LUB POST month not verified: {wanted}")
def fetch_lub(s,d,v,r,y,m):
    html=lub_html(s,y,m); h,row=find_depot(html,d,v,["TOTAL LUB KMPL"])
    if not row:return {}
    # Prefer explicit monthly/cumulative Total Lub KMPL columns. Never calculate from oil quantities.
    month_i=None; upto_i=None; total_cols=[]
    for i,x in enumerate(h):
        hx=n(x)
        if "TOTAL LUB KMPL" not in hx: continue
        total_cols.append(i)
        if "FOR" in hx: month_i=i
        if "UP TO" in hx or "UPTO" in hx: upto_i=i
    if month_i is None and total_cols: month_i=total_cols[0]
    if upto_i is None and len(total_cols)>1: upto_i=total_cols[-1]
    return {"TOTAL LUB KMPL":{"month":at(row,month_i),"upto":at(row,upto_i)}}

def tyre_row(h,rows,code):
    di=core.col(h,["DEPOT","DCP_CODE","DCP CODE"]); si=core.col(h,["TYRE SIZE","SIZE"])
    for row in rows:
        if di is not None and si is not None and di<len(row) and si<len(row) and n(row[di])==n(code) and n(row[si])=="ALL TYRE SIZES TOTAL": return row
    return None
def tyre_values(h,row):
    if not row:return {}
    rv=lambda a:core.row_value(h,row,a)
    return {"N.T.S RATE":rv(["NEW TYRE %","NEW %"]),"TTL SCP Rate":rv(["TOTAL %"]),"RC TYRE LIFE":rv(["RC_MILEAGE","RC MILEAGE"]),"AVG TYRE LIFE":rv(["AVG_TOTAL MILEAGE","AVG TOTAL MILEAGE"]),"Ist RC S Rate":rv(["IST RC %"]),"RT Factor":rv(["RT_FACTOR","RT FACTOR"]),"NEW TYRE LIFE":rv(["NEW MILEAGE"])}
def tyre_page(s,path,d,y,m):
    info=TYRE_SITE.get(n(d));
    if not info: raise RuntimeError(f"No tyre mapping for {d}")
    code,zone,region=info
    html=get_html(s,core.TYRE_BASE,path,{"zone":zone,"region":region,"depot":code,"month_year":datetime(y,m,1).strftime("%b-%Y").upper(),"tyre_size":"All Tyre Sizes Total"})
    h,rows=core.find_table(html,["DEPOT"],["NEW MILEAGE","AVG TOTAL MILEAGE","RT"])
    return h,tyre_row(h,rows,code)

def normalize_tyre(dct):
    out={}
    for k,v in dct.items():
        if k in {"AVG TYRE LIFE","NEW TYRE LIFE","RC TYRE LIFE"} and v is not None: v=v/100000.0
        out[k]=v
    return out

def pdf_headers_and_row(table,d,code):
    rows=[[str(c or "").strip() for c in row] for row in table if row]
    if not rows:return None,None
    target=None
    for ri,row in enumerate(rows):
        txt=" | ".join(row)
        if (n(code) in n(txt) or n(d) in n(txt)) and ("ALL TYRE SIZES TOTAL" in n(txt) or len(row)>=7): target=ri; break
    if target is None:return None,None
    width=max(len(x) for x in rows[:target+1]); hdr=[]
    for c in range(width):
        parts=[]
        for rr in rows[max(0,target-5):target]:
            if c<len(rr) and rr[c] and n(rr[c]) not in parts: parts.append(n(rr[c]))
        hdr.append(" | ".join(parts))
    row=rows[target]+[""]*(width-len(rows[target]))
    return hdr,row
def tyre_pdf_fallback(s,d,y,m):
    # APSRTC web tyre tables do not expose FY2024-25 reliably. Use the official monthly TRS booklet only for that FY.
    fy=selected_fy(y,m)
    if fy!="2024-25": return {},{}
    mon=datetime(y,m,1).strftime("%b").lower(); url=f"{core.MED_BASE}/trs_booklet/2024-25/{mon}-{y}.pdf"
    rr=s.get(url,timeout=60); rr.raise_for_status()
    if not rr.content.startswith(b"%PDF"): raise RuntimeError("TRS booklet is not a PDF")
    code=TYRE_SITE[n(d)][0]; monthly={}; upto={}
    with pdfplumber.open(io.BytesIO(rr.content)) as pdf:
        for page in pdf.pages:
            ptxt=n(page.extract_text() or "")
            for table in page.extract_tables() or []:
                h,row=pdf_headers_and_row(table,d,code)
                if not row: continue
                vals=normalize_tyre(tyre_values(h,row))
                if not any(v is not None for v in vals.values()): continue
                if "STATEMENT D" in ptxt or "FOR THE MONTH" in ptxt: monthly.update({k:v for k,v in vals.items() if v is not None})
                if "STATEMENT E" in ptxt or "UP TO THE MONTH" in ptxt or "UPTO THE MONTH" in ptxt: upto.update({k:v for k,v in vals.items() if v is not None})
    if monthly or upto: print(f"TYRE PDF fallback used: {url}")
    return monthly,upto
def fetch_tyre(s,d,y,m,need_upto):
    dm={}; eu={}
    try:
        hd,rd=tyre_page(s,"d_statement_final.php",d,y,m); dm=normalize_tyre(tyre_values(hd,rd))
        if need_upto:
            he,re=tyre_page(s,"e_statement_final.php",d,y,m); eu=normalize_tyre(tyre_values(he,re))
    except Exception as exc: print(f"TYRE WEB {y:04d}-{m:02d}: {exc}")
    if selected_fy(y,m)=="2024-25" and (not any(v is not None for v in dm.values()) or (need_upto and not any(v is not None for v in eu.values()))):
        pm,pu=tyre_pdf_fallback(s,d,y,m); dm={**pm,**{k:v for k,v in dm.items() if v is not None}}; eu={**pu,**{k:v for k,v in eu.items() if v is not None}}
    return {name:{"month":dm.get(name),"upto":eu.get(name) if need_upto else None} for name in TYRE}

def fetch_month(s,d,v,r,y,m,need_upto):
    result={}; jobs=[("HSD",lambda:fetch_hsd(s,d,v,r,y,m)),("PRODUCT",lambda:fetch_dimension(s,"prodkmpl_um.php","PRODUCT","PRODUCT",d,r,y,m)),("ENGINE",lambda:fetch_dimension(s,"engkmpl_um.php","ENGINE TYPE","ENGINE",d,r,y,m)),("LUB",lambda:fetch_lub(s,d,v,r,y,m)),("BD",lambda:fetch_bd(s,d,v,r,y,m)),("MED",lambda:fetch_med(s,d,v,r,y,m)),("SPRING",lambda:fetch_spring(s,d,v,r,y,m)),("TYRE",lambda:fetch_tyre(s,d,y,m,need_upto))]
    for label,fn in jobs:
        try: result.update(fn())
        except Exception as exc: print(f"{label} {y:04d}-{m:02d}: {MANUAL}/BLANK: {exc}")
    return result

def new_store(fys): return {"fys":fys,"order":[],"rows":{}}
def ensure(st,name):
    if name not in st["rows"]: st["rows"][name]={fy:{"months":{},"upto":None} for fy in st["fys"]}; st["order"].append(name)
def seed(st):
    for x in FIXED: ensure(st,x)
def valid_dynamic(name):
    if not (name.startswith("PRODUCT:") or name.startswith("ENGINE:")): return True
    raw=name.split(":",1)[1].strip(); return bool(re.search(r"[A-Za-z]",raw)) and n(raw) not in {"TOTAL","GRAND TOTAL"}
def load_existing(values,fys):
    st=new_store(fys); current=None
    for row in values[1:] if values else []:
        if row and str(row[0]).strip():
            candidate=str(row[0]).strip(); current=candidate if valid_dynamic(candidate) else None
            if current: ensure(st,current)
        if not current or len(row)<2 or str(row[1]).strip() not in fys: continue
        fy=str(row[1]).strip()
        for i,mon in enumerate(MONTHS):
            c=2+i
            if c<len(row) and str(row[c]).strip()!="": st["rows"][current][fy]["months"][mon]=row[c]
        if len(row)>15 and str(row[15]).strip()!="": st["rows"][current][fy]["upto"]=row[15]
    seed(st); return st
def apply(st,fy,y,m,data,need_upto,repair):
    mon=month_name(m); returned=set(data)
    for name,v in data.items():
        if not valid_dynamic(name): continue
        ensure(st,name); dynamic=name.startswith("PRODUCT:") or name.startswith("ENGINE:")
        if v.get("month") is not None: st["rows"][name][fy]["months"][mon]=v["month"]
        elif repair and not dynamic: st["rows"][name][fy]["months"].setdefault(mon,MANUAL)
        if need_upto and v.get("upto") is not None: st["rows"][name][fy]["upto"]=v["upto"]
        elif need_upto and repair and not dynamic and st["rows"][name][fy]["upto"] is None: st["rows"][name][fy]["upto"]=MANUAL
    if repair:
        for name in FIXED:
            ensure(st,name)
            if name not in returned:
                st["rows"][name][fy]["months"].setdefault(mon,MANUAL)
                if need_upto and st["rows"][name][fy]["upto"] is None: st["rows"][name][fy]["upto"]=MANUAL
def reorder(st):
    p=[x for x in st["order"] if x.startswith("PRODUCT:") and valid_dynamic(x)]; e=[x for x in st["order"] if x.startswith("ENGINE:") and valid_dynamic(x)]
    tail=FIXED[2:]; base=FIXED[:2]+p+e+tail; st["order"]=base+[x for x in st["order"] if x not in base and valid_dynamic(x)]
def matrix(st):
    reorder(st); out=[HEADERS]
    for name in st["order"]:
        for j,fy in enumerate(st["fys"]):
            v=st["rows"][name][fy]; out.append([name if j==0 else "",fy]+[v["months"].get(mon,"") for mon in MONTHS]+["",v["upto"] if v["upto"] is not None else ""])
    return out

def make_xlsx(display,mat,fys):
    path=REPORTS/f"{display}_ANNUAL_KPI_DASHBOARD.xlsx"; wb=Workbook(); ws=wb.active; ws.title=SHEET_TITLE
    for r in mat: ws.append(r)
    thin=Side(style="thin"); border=Border(left=thin,right=thin,top=thin,bottom=thin); fill=PatternFill("solid",fgColor="1F4E78"); colors=["1F4E79","008000","C00000"]
    for c in range(1,17): ws.cell(1,c).font=Font(bold=True,color="FFFFFF"); ws.cell(1,c).fill=fill; ws.cell(1,c).alignment=Alignment(horizontal="center")
    for r in range(2,ws.max_row+1):
        fy=str(ws.cell(r,2).value or ""); color=colors[fys.index(fy)] if fy in fys else "000000"
        for c in range(1,17): ws.cell(r,c).border=border; ws.cell(r,c).alignment=Alignment(horizontal="center",vertical="center"); ws.cell(r,c).font=Font(color=color)
    for r in range(2,ws.max_row+1,3): ws.merge_cells(start_row=r,start_column=1,end_row=min(r+2,ws.max_row),end_column=1); ws.cell(r,1).font=Font(bold=True,color="000000")
    for c,w in {1:28,2:12,15:3,16:12}.items(): ws.column_dimensions[get_column_letter(c)].width=w
    for c in range(3,15): ws.column_dimensions[get_column_letter(c)].width=11
    ws.freeze_panes="C2"; wb.save(path); return path

def format_sheet(spreadsheet_id,mat,fys):
    sid=sheet_id(spreadsheet_id,SHEET_TITLE); svc=sheets_service(); maxr=max(1000,len(mat)+20)
    svc.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id,body={"requests":[{"unmergeCells":{"range":{"sheetId":sid,"startRowIndex":1,"endRowIndex":maxr,"startColumnIndex":0,"endColumnIndex":1}}}]}).execute()
    write_values(spreadsheet_id,f"'{SHEET_TITLE}'!A1",mat)
    req=[]; colors=[{"red":.12,"green":.31,"blue":.48},{"red":0,"green":.5,"blue":0},{"red":.75,"green":0,"blue":0}]
    req.append({"repeatCell":{"range":{"sheetId":sid,"startRowIndex":0,"endRowIndex":1,"startColumnIndex":0,"endColumnIndex":16},"cell":{"userEnteredFormat":{"backgroundColor":{"red":.12,"green":.31,"blue":.47},"textFormat":{"bold":True,"foregroundColor":{"red":1,"green":1,"blue":1}},"horizontalAlignment":"CENTER"}},"fields":"userEnteredFormat"}})
    for r in range(1,len(mat)):
        fy=str(mat[r][1]); color=colors[fys.index(fy)]
        req.append({"repeatCell":{"range":{"sheetId":sid,"startRowIndex":r,"endRowIndex":r+1,"startColumnIndex":1,"endColumnIndex":16},"cell":{"userEnteredFormat":{"textFormat":{"foregroundColor":color},"horizontalAlignment":"CENTER","verticalAlignment":"MIDDLE"}},"fields":"userEnteredFormat(textFormat.foregroundColor,horizontalAlignment,verticalAlignment)"}})
    for r in range(1,len(mat),3):
        end=min(r+3,len(mat)); req.append({"mergeCells":{"range":{"sheetId":sid,"startRowIndex":r,"endRowIndex":end,"startColumnIndex":0,"endColumnIndex":1},"mergeType":"MERGE_ALL"}})
        req.append({"repeatCell":{"range":{"sheetId":sid,"startRowIndex":r,"endRowIndex":end,"startColumnIndex":0,"endColumnIndex":1},"cell":{"userEnteredFormat":{"textFormat":{"bold":True,"foregroundColor":{"red":0,"green":0,"blue":0}},"verticalAlignment":"MIDDLE"}},"fields":"userEnteredFormat(textFormat,verticalAlignment)"}})
    req += [{"updateDimensionProperties":{"range":{"sheetId":sid,"dimension":"COLUMNS","startIndex":0,"endIndex":1},"properties":{"pixelSize":240},"fields":"pixelSize"}},{"updateDimensionProperties":{"range":{"sheetId":sid,"dimension":"COLUMNS","startIndex":1,"endIndex":2},"properties":{"pixelSize":95},"fields":"pixelSize"}},{"updateDimensionProperties":{"range":{"sheetId":sid,"dimension":"COLUMNS","startIndex":2,"endIndex":14},"properties":{"pixelSize":85},"fields":"pixelSize"}},{"updateDimensionProperties":{"range":{"sheetId":sid,"dimension":"COLUMNS","startIndex":14,"endIndex":15},"properties":{"pixelSize":20},"fields":"pixelSize"}},{"updateDimensionProperties":{"range":{"sheetId":sid,"dimension":"COLUMNS","startIndex":15,"endIndex":16},"properties":{"pixelSize":95},"fields":"pixelSize"}},{"updateSheetProperties":{"properties":{"sheetId":sid,"gridProperties":{"frozenRowCount":1,"frozenColumnCount":2}},"fields":"gridProperties.frozenRowCount,gridProperties.frozenColumnCount"}}]
    svc.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id,body={"requests":req}).execute()

def meta_version(spreadsheet_id):
    try:
        vals=read_values(spreadsheet_id,f"'{META_TITLE}'!A:B")
        d={str(r[0]):str(r[1]) for r in vals[1:] if len(r)>=2}; return d.get("LAYOUT_VERSION","")
    except Exception:return ""
def main():
    p=argparse.ArgumentParser(); p.add_argument("--depot",required=True); p.add_argument("--selected-month",required=True); p.add_argument("--years",default=""); a=p.parse_args()
    sy,sm=map(int,a.selected_month.split("-")); datetime(sy,sm,1); fys=fy_triplet(sy,sm)
    vehicle,display,region=core.depot_info(a.depot); folder=os.getenv("KPI_DRIVE_FOLDER_ID",core.DEFAULT_DRIVE_FOLDER); name=f"{display}_ANNUAL_KPI_DASHBOARD"
    existing=find_file(folder,name); repair=(not existing) or meta_version(existing["id"])!="3"
    st=new_store(fys) if not existing else load_existing(read_values(existing["id"],f"'{SHEET_TITLE}'!A:P"),fys); seed(st); s=login()
    if repair:
        print(f"FULL REPAIR/BUILD {display}: {', '.join(fys)}")
        for fy in fys:
            months=fy_months(fy,sy,sm)
            for i,(y,m) in enumerate(months):
                need=i==len(months)-1; print(f"FETCH {fy} {y:04d}-{m:02d}{' + Upto' if need else ''}"); apply(st,fy,y,m,fetch_month(s,display,vehicle,region,y,m,need),need,True)
    else:
        fy=selected_fy(sy,sm); print(f"INCREMENTAL {display} {a.selected_month}"); apply(st,fy,sy,sm,fetch_month(s,display,vehicle,region,sy,sm,True),True,False)
    mat=matrix(st); xlsx=make_xlsx(display,mat,fys)
    if not existing:
        up=upload_xlsx_as_google_sheet(xlsx,folder,name); sid=up["id"]; link=up.get("webViewLink","")
    else: sid=existing["id"]; link=existing.get("webViewLink","")
    format_sheet(sid,mat,fys); ensure_hidden_sheet(sid,META_TITLE); write_values(sid,f"'{META_TITLE}'!A1",[["KEY","VALUE"],["DEPOT",display],["FYS",','.join(fys)],["LAST_SELECTED_MONTH",a.selected_month],["LAYOUT_VERSION","3"]])
    print(f"ANNUAL_KPI_DASHBOARD_SUCCESS: {link}"); print(f"GOOGLE_SHEET_ID: {sid}"); return 0
if __name__=="__main__":
    try: sys.exit(main())
    except Exception as exc: print(f"ANNUAL_KPI_FAILURE: {exc}"); raise
