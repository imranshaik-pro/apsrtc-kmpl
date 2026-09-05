#!/usr/bin/env python3
"""Generate a monthly vehicle HSD KMPL report and upload it as a Google Sheet."""

import argparse
import json
import os
import sys
from calendar import monthrange
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.auth.client import login
from src.integrations.google_drive import upload_xlsx_as_google_sheet
from src.parser.vehicle_parser import parse_vehicle_rows

PROJECT_DIR = Path(__file__).resolve().parent
MAPPING_FILE = PROJECT_DIR / "depot_mapping.json"
DEFAULT_GDRIVE_FOLDER = "1O6rKH39INogYxsNI9IepEO9UJq1MMnPj"
IST = ZoneInfo("Asia/Kolkata")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def get_style(value):
    if value is None or pd.isna(value):
        return None, None
    if value <= 5.00:
        return PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"), Font(color="FFFFFF", bold=True)
    if value <= 5.10:
        return PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"), Font(color="FFFFFF", bold=True)
    if value <= 5.20:
        return PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"), Font(color="000000")
    if value <= 5.30:
        return PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"), Font(color="000000")
    return PatternFill(start_color="008000", end_color="008000", fill_type="solid"), Font(color="FFFFFF", bold=True)


def load_depot_mapping():
    with MAPPING_FILE.open("r", encoding="utf-8") as f:
        return json.load(f)


def get_depot_info(depot_key: str):
    mapping = load_depot_mapping()
    for key in [depot_key, depot_key.lower()]:
        if key in mapping:
            info = mapping[key]
            return info["vehicle_depot"], info.get("display_name", key.upper())
    for info in mapping.values():
        if str(info.get("display_name", "")).upper() == depot_key.upper():
            return info["vehicle_depot"], info.get("display_name", depot_key.upper())
    raise ValueError(f"Depot '{depot_key}' not found.")


def consolidate_day_records(records):
    consolidated = {}
    for rec in records:
        vehicle_no = rec["vehicle_no"]
        if vehicle_no not in consolidated:
            consolidated[vehicle_no] = {
                "total_kms": rec["for_day_total_kms"],
                "hsd": rec["for_day_hsd"],
                "operation_type": rec.get("operation_type", ""),
                "engine_type": rec.get("engine_type", ""),
                "up_to_day_kmpl": rec.get("up_to_day_kmpl"),
            }
        else:
            consolidated[vehicle_no]["total_kms"] += rec["for_day_total_kms"]
            consolidated[vehicle_no]["hsd"] += rec["for_day_hsd"]
            if not consolidated[vehicle_no]["up_to_day_kmpl"] and rec.get("up_to_day_kmpl"):
                consolidated[vehicle_no]["up_to_day_kmpl"] = rec.get("up_to_day_kmpl")

    for data in consolidated.values():
        hsd = data["hsd"]
        data["kmpl"] = (data["total_kms"] / hsd) if hsd and hsd > 0 else None
    return consolidated


def fetch_daily_data(session, date_obj, vehicle_depot):
    response = session.post(
        "http://103.44.14.20/med/vehkmpl.php",
        data={"fyymm": date_obj.strftime("%d/%m/%Y"), "dept": vehicle_depot},
        timeout=30,
    )
    response.raise_for_status()
    records = parse_vehicle_rows(response.text)
    return consolidate_day_records(records) if records else {}


def apply_formatting(workbook):
    ws = workbook.active
    ws.freeze_panes = "E2"

    # Columns E onward contain daily KMPL and the month-end KMPL.
    # Normalize every nonblank value to a real numeric cell and force two-decimal
    # display so 5 -> 5.00 and 5.1 -> 5.10 after conversion to Google Sheets.
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=ws.max_column):
        for cell in row:
            if cell.value in (None, ""):
                continue
            try:
                numeric_value = float(str(cell.value).replace(",", "").strip())
            except (TypeError, ValueError):
                continue

            cell.value = round(numeric_value, 2)
            cell.number_format = "0.00"
            fill, font = get_style(numeric_value)
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            cell.border = THIN_BORDER

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 30)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--depot", required=True)
    parser.add_argument("--month", required=True, help="YYYY-MM")
    args = parser.parse_args()

    try:
        year, month = map(int, args.month.split("-"))
        start_date = datetime(year, month, 1)
        _, days_in_month = monthrange(year, month)
    except ValueError:
        print("INVALID_MONTH: Use YYYY-MM")
        return 2

    now_ist = datetime.now(IST)
    selected_month = (year, month)
    current_month = (now_ist.year, now_ist.month)
    if selected_month > current_month:
        print("INVALID_MONTH: LOL! Can't retrieve future-month data.")
        return 2

    last_day = days_in_month
    if selected_month == current_month:
        last_day = now_ist.day - 1
        if last_day < 1:
            print("NO_DATA: No completed day exists yet this month.")
            return 1

    vehicle_depot, display_name = get_depot_info(args.depot)
    print(f"Monthly report: {display_name} {args.month}; days 1-{last_day}")

    session = login()
    vehicle_data = {}
    for day_num in range(1, last_day + 1):
        date_obj = datetime(year, month, day_num)
        print(f"Fetching {date_obj.strftime('%d/%m/%Y')}...")
        day_records = fetch_daily_data(session, date_obj, vehicle_depot)
        for vehicle_no, data in day_records.items():
            if vehicle_no not in vehicle_data:
                vehicle_data[vehicle_no] = {
                    "op_type": data.get("operation_type", ""),
                    "engine": data.get("engine_type", ""),
                    "days": {},
                    "up_to_day_latest": None,
                }
            vehicle_data[vehicle_no]["days"][day_num] = data.get("kmpl")
            if data.get("up_to_day_kmpl") is not None:
                vehicle_data[vehicle_no]["up_to_day_latest"] = data.get("up_to_day_kmpl")

    if not vehicle_data:
        print("NO_DATA: No vehicle data found.")
        return 1

    rows = []
    for idx, (vehicle_no, data) in enumerate(sorted(vehicle_data.items()), 1):
        row = {
            "SL No": idx,
            "Vehicle No": vehicle_no,
            "Op Type": data["op_type"],
            "Engine Type": data["engine"],
        }
        for day_num in range(1, days_in_month + 1):
            value = data["days"].get(day_num)
            row[str(day_num)] = value if value is not None else ""
        row["Up-To-Day (Month End)"] = data["up_to_day_latest"] if data["up_to_day_latest"] is not None else ""
        rows.append(row)

    reports_dir = PROJECT_DIR / "reports"
    reports_dir.mkdir(exist_ok=True)
    xlsx_name = f"{display_name}_{args.month}.xlsx"
    xlsx_path = reports_dir / xlsx_name

    df = pd.DataFrame(rows)
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Monthly KMPL", index=False)

    workbook = load_workbook(xlsx_path)
    apply_formatting(workbook)
    workbook.save(xlsx_path)

    folder_id = os.getenv("MONTHLY_DRIVE_FOLDER_ID", DEFAULT_GDRIVE_FOLDER)
    target_sheet_name = f"{display_name}_{args.month}"
    uploaded = upload_xlsx_as_google_sheet(
        xlsx_path, folder_id=folder_id, sheet_name=target_sheet_name
    )

    if uploaded.get("already_existed"):
        print(f"ALREADY_DELIVERED: {uploaded.get('webViewLink')}")
    else:
        print(f"MONTHLY_AUTOMATION_SUCCESS: {uploaded.get('webViewLink')}")
    print(f"GOOGLE_SHEET_ID: {uploaded.get('id')}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
