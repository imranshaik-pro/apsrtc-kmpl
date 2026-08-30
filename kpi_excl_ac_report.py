#!/usr/bin/env python3
"""
KPI Report for HSD KMPL (EXCL AC) using daily vehicle data.
Includes full financial year months (Apr to Mar).
"""

import os
import sys
import json
import re
import subprocess
import tempfile
from datetime import datetime
from calendar import monthrange
from typing import Dict, List, Optional, Tuple
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from src.auth.client import login
from src.parser.vehicle_parser import parse_vehicle_rows
from src.parser.region_parser import parse_region_html

PROJECT_DIR = "/home/imran/apsrtc-kmpl"
MAPPING_FILE = os.path.join(PROJECT_DIR, "depot_mapping.json")
GDRIVE_FOLDER_KPI = "1uQlJJcrv7TbKAXtkFtrrCqZac31aWqL9"
VEHICLE_TYPE_MAPPING = os.path.join(PROJECT_DIR, "vehicle_type_mapping.json")

MONTH_NAMES = {
    4: "Apr", 5: "May", 6: "Jun", 7: "Jul",
    8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov",
    12: "Dec", 1: "Jan", 2: "Feb", 3: "Mar"
}

def load_depot_mapping():
    with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def get_depot_info(depot_key: str):
    mapping = load_depot_mapping()
    for key in [depot_key, depot_key.lower()]:
        if key in mapping:
            info = mapping[key]
            return info["vehicle_depot"], info.get("display_name", key.upper()), info.get("region_code")
    raise ValueError(f"Depot '{depot_key}' not found.")

def load_type_mapping():
    with open(VEHICLE_TYPE_MAPPING, 'r', encoding='utf-8') as f:
        return json.load(f)

def fetch_daily_data(session, date_obj, vehicle_depot):
    """Fetch and consolidate vehicle data for a single day."""
    day_str = date_obj.strftime("%d/%m/%Y")
    response = session.post(
        "http://103.44.14.20/med/vehkmpl.php",
        data={"fyymm": day_str, "dept": vehicle_depot},
        timeout=30
    )
    response.raise_for_status()
    records = parse_vehicle_rows(response.text)
    if not records:
        return {}
    consolidated = {}
    for rec in records:
        v = rec["vehicle_no"]
        if v not in consolidated:
            consolidated[v] = {
                "total_kms": rec["for_day_total_kms"],
                "hsd": rec["for_day_hsd"],
                "op_type": rec.get("operation_type", ""),
            }
        else:
            consolidated[v]["total_kms"] += rec["for_day_total_kms"]
            consolidated[v]["hsd"] += rec["for_day_hsd"]
    return consolidated

def compute_monthly_excl_ac(session, year, month, vehicle_depot):
    """
    Compute cumulative EXCL AC KMPL for the month by aggregating daily data.
    Returns: (total_kms, total_hsd, kmpl) for NAC vehicles.
    """
    mapping = load_type_mapping()
    _, last_day = monthrange(year, month)
    total_kms = 0.0
    total_hsd = 0.0
    for day in range(1, last_day + 1):
        date_obj = datetime(year, month, day)
        day_data = fetch_daily_data(session, date_obj, vehicle_depot)
        for v, rec in day_data.items():
            op = rec["op_type"].strip().upper()
            code = op.split()[0] if op else ""
            # Skip AC vehicles
            if mapping.get(code) == "AC":
                continue
            total_kms += rec["total_kms"]
            total_hsd += rec["hsd"]
    kmpl = total_kms / total_hsd if total_hsd > 0 else None
    # Debug print for Jan, Feb, Mar
    if month in [1,2,3]:
        print(f"DEBUG {year}-{month:02d}: total_kms={total_kms:.2f}, total_hsd={total_hsd:.2f}, kmpl={kmpl}")
    return total_kms, total_hsd, kmpl

def fetch_region_target(session, region_code, depot_display, year, month):
    """Fetch target from region report for the first day of the month."""
    date_obj = datetime(year, month, 1)
    date_str = date_obj.strftime("%Y-%m-%d")
    url = "http://103.44.14.20/med/achsd1.php"
    params = {"action": "", "fdate": date_str, "rreg": region_code}
    response = session.get(url, params=params, timeout=30)
    response.raise_for_status()
    records = parse_region_html(response.text)
    for rec in records:
        if rec.get("depot") == depot_display and rec.get("category") == "NAC":
            return rec.get("target_kmpl")
    return None

def ensure_gdrive_account():
    result = subprocess.run(["gdrive", "account", "list"], stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True)
    if "iamrebel1984@gmail.com" not in result.stdout:
        subprocess.run(["gdrive", "account", "switch", "iamrebel1984@gmail.com"], check=True)

def upload_to_drive(file_path):
    ensure_gdrive_account()
    cmd = ["gdrive", "files", "upload", "--parent", GDRIVE_FOLDER_KPI, file_path]
    result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True)
    if result.returncode != 0:
        raise RuntimeError(f"Upload failed: {result.stderr}")
    return result.stdout

def generate_excel(kpi_data, depot, selected_month):
    rows = []
    # Group by financial year
    years = {}
    for month, data in kpi_data.items():
        y = int(month[:4])
        m = int(month[5:])
        if m >= 4:
            fy = f"{y}-{y+1}"
        else:
            fy = f"{y-1}-{y}"
        years.setdefault(fy, []).append(month)

    month_order = [4,5,6,7,8,9,10,11,12,1,2,3]
    for fy, months in sorted(years.items()):
        row = {"Financial Year": fy, "Target": None}
        # Target is same for all months; pick from first month
        for m in months:
            if kpi_data[m].get("target") is not None:
                row["Target"] = kpi_data[m]["target"]
                break
        # Map month numbers to values
        month_map = {int(m[5:]): m for m in months}
        for m in month_order:
            col_name = MONTH_NAMES.get(m, str(m))
            if m in month_map:
                month_str = month_map[m]
                val = kpi_data[month_str].get("kmpl")
                row[col_name] = val
            else:
                row[col_name] = None
        # Cumulative: the value of the last month in this financial year (Mar if available, else last)
        last_month = None
        for m in reversed(month_order):
            if m in month_map:
                last_month = month_map[m]
                break
        if last_month:
            row["Cumulative (Up to Month)"] = kpi_data[last_month].get("kmpl")
        else:
            row["Cumulative (Up to Month)"] = None
        rows.append(row)

    df = pd.DataFrame(rows)
    df = df.round(2)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    with pd.ExcelWriter(tmp_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="HSD KMPL EXCL AC", index=False)

    wb = load_workbook(tmp_path)
    ws = wb.active
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 30)
    wb.save(tmp_path)
    return tmp_path

def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--depot", required=True)
    parser.add_argument("--month", required=True)  # YYYY-MM
    args = parser.parse_args()

    vehicle_depot, display_name, region_code = get_depot_info(args.depot)
    print(f"📊 Generating HSD KMPL (EXCL AC) report for {display_name} up to {args.month}")

    selected_year = int(args.month[:4])
    selected_month = int(args.month[5:])

    session = login()

    # Determine financial years to include: 2024-25, 2025-26, and 2026-27 (partial up to selected month)
    # We'll generate list of all months from April 2024 to selected month, including Jan-Feb-Mar.
    months = []
    start_year = 2024
    # For each year from start_year to selected_year
    for y in range(start_year, selected_year + 1):
        # If this is the selected year, we only go up to selected_month
        end_month = selected_month if y == selected_year else 12
        # For each month from 4 to end_month
        for m in range(4, end_month + 1):
            months.append(f"{y}-{m:02d}")
        # If this is the selected year and selected_month < 3, we also need to include Jan-Mar of next year? No, because selected year is the year of the selected month.
        # Actually, if selected_month is in Jan-Mar, we need to include those months of the same year.
        # But our loop for y == selected_year will include months 4 to selected_month (if selected_month >=4), else it will add nothing.
        # We need to handle the case where selected_month < 4: then we should include months from 4 to 12 of previous year, and 1 to selected_month of current year.
        # Our current loop doesn't handle that. Let's redo:
        # We'll use a while loop to add months sequentially.
    # Simpler: we'll just use a while loop to add months from April 2024 to selected month.
    # We'll define a date and increment month by month.
    current = datetime(2024, 4, 1)
    end = datetime(selected_year, selected_month, 1)
    months = []
    while current <= end:
        months.append(current.strftime("%Y-%m"))
        # Move to next month
        if current.month == 12:
            current = current.replace(year=current.year+1, month=1)
        else:
            current = current.replace(month=current.month+1)
    # Now months contains all months from Apr 2024 to selected month (including Jan, Feb, Mar).

    print(f"📆 Months to process: {len(months)}")

    kpi_data = {}

    for month in months:
        y, m = map(int, month.split("-"))
        print(f"Computing EXCL AC for {month}...")
        total_kms, total_hsd, kmpl = compute_monthly_excl_ac(session, y, m, vehicle_depot)
        if kmpl is not None:
            # Fetch target for this month (use region report for the first day)
            target = fetch_region_target(session, region_code, display_name, y, m)
            kpi_data[month] = {"kmpl": kmpl, "target": target}
        else:
            print(f"  No data for {month}")

    if not kpi_data:
        print("No data found for any month.")
        sys.exit(1)

    excel_path = generate_excel(kpi_data, display_name, args.month)

    filename = f"HSD_KMPL_EXCL_AC_{display_name}_{args.month.replace('-','_')}.xlsx"
    final_path = os.path.join(os.path.dirname(excel_path), filename)
    os.rename(excel_path, final_path)

    print("⬆️ Uploading to Google Drive...")
    upload_output = upload_to_drive(final_path)

    match = re.search(r'ViewUrl:\s*(https://[^\s]+)', upload_output)
    if match:
        print(f"✅ Upload successful!\n🔗 {match.group(1)}")
    else:
        print(f"✅ Upload successful!\n📁 Folder: https://drive.google.com/drive/folders/{GDRIVE_FOLDER_KPI}")

    os.unlink(final_path)
    print("🎉 Done!")

if __name__ == "__main__":
    main()
