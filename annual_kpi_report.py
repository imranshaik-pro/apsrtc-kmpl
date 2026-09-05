#!/usr/bin/env python3
"""Generate an APSRTC annual KPI summary and upload it as a Google Sheet.

Business rules:
- Financial year is April-March.
- Completed FY uses March's cumulative (Up To the Month CY) source value.
- Current FY uses the latest completed month.
- Product and Engine rows are preserved exactly as APSRTC returns them; no merging.
- TOTAL LUB KMPL is read directly and displayed with no decimals.
- Spring Consumption per Lakh Kms is read directly and displayed with 2 decimals.
- Tyre life values are divided by 100000 and displayed with 2 decimals.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from calendar import monthrange
from datetime import datetime
from pathlib import Path
from typing import Iterable
from zoneinfo import ZoneInfo

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.auth.client import login
from src.integrations.google_drive import upload_xlsx_as_google_sheet

PROJECT_DIR = Path(__file__).resolve().parent
MAPPING_FILE = PROJECT_DIR / "depot_mapping.json"
IST = ZoneInfo("Asia/Kolkata")
DEFAULT_DRIVE_FOLDER = "1uQlJJcrv7TbKAXtkFtrrCqZac31aWqL9"

MED_BASE = "http://103.44.14.20/med"
MEDNEW_BASE = "http://103.44.14.20/mednew"
TYRE_BASE = "http://103.44.14.20/tyres"

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def norm(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip().upper()


def number(value):
    if value is None:
        return None
    text = str(value).strip().replace(",", "").replace("%", "").replace("−", "-").replace("–", "-")
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else None


def load_mapping():
    with MAPPING_FILE.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def depot_info(depot_key: str):
    wanted = norm(depot_key)
    for key, info in load_mapping().items():
        if norm(key) == wanted or norm(info.get("display_name")) == wanted:
            return (
                info["vehicle_depot"],
                info.get("display_name", key.upper()),
                info.get("region_code", ""),
            )
    raise ValueError(f"Depot '{depot_key}' not found in depot_mapping.json")


def parse_fy(fy: str):
    match = re.fullmatch(r"(20\d{2})-(\d{2})", fy.strip())
    if not match:
        raise ValueError(f"Invalid financial year '{fy}'. Use YYYY-YY, for example 2025-26.")
    start = int(match.group(1))
    expected = (start + 1) % 100
    if int(match.group(2)) != expected:
        raise ValueError(f"Invalid financial year '{fy}'.")
    return start, start + 1


def current_fy(now=None):
    now = now or datetime.now(IST)
    start = now.year if now.month >= 4 else now.year - 1
    return f"{start}-{str(start + 1)[-2:]}"


def effective_month_for_fy(fy: str, now=None):
    now = now or datetime.now(IST)
    start, end = parse_fy(fy)
    if fy > current_fy(now):
        raise ValueError("LOL! Can't retrieve future financial-year data.")
    if fy < current_fy(now):
        return end, 3
    y, m = now.year, now.month - 1
    if m == 0:
        y, m = now.year - 1, 12
    if (y, m) < (start, 4):
        raise ValueError(f"No completed month exists yet for FY {fy}.")
    return y, m


def request_html(session, base: str, path: str, params: dict):
    url = f"{base}/{path}"
    last = None
    for method in ("get", "post"):
        try:
            response = getattr(session, method)(
                url,
                params=params if method == "get" else None,
                data=params if method == "post" else None,
                timeout=45,
            )
            response.raise_for_status()
            if response.text and "<table" in response.text.lower():
                return response.text
            last = RuntimeError(f"{url} returned no table")
        except Exception as exc:
            last = exc
    raise RuntimeError(f"Unable to load {url}: {last}")


def table_matrix(table):
    rows = []
    for tr in table.find_all("tr"):
        cells = tr.find_all(["th", "td"], recursive=False)
        if cells:
            rows.append([c.get_text(" ", strip=True) for c in cells])
    return rows


def expanded_headers(table):
    trs = table.find_all("tr")
    header_trs = []
    for tr in trs:
        if tr.find_all("th", recursive=False):
            header_trs.append(tr)
        else:
            break
    if not header_trs:
        return [], table_matrix(table)

    grid, active = [], {}
    for r_idx, tr in enumerate(header_trs):
        row, col_idx = [], 0

        def put(idx, text):
            while len(row) <= idx:
                row.append("")
            row[idx] = text

        for cell in tr.find_all("th", recursive=False):
            while col_idx in active and active[col_idx][1] > r_idx:
                put(col_idx, active[col_idx][0])
                col_idx += 1
            text = cell.get_text(" ", strip=True)
            rs = int(cell.get("rowspan", 1) or 1)
            cs = int(cell.get("colspan", 1) or 1)
            for j in range(cs):
                put(col_idx + j, text)
                if rs > 1:
                    active[col_idx + j] = (text, r_idx + rs)
            col_idx += cs
        for idx, (text, until) in active.items():
            if until > r_idx:
                put(idx, text)
        grid.append(row)

    width = max((len(r) for r in grid), default=0)
    for row in grid:
        row.extend([""] * (width - len(row)))
    headers = []
    for c in range(width):
        parts = []
        for r in range(len(grid)):
            part = norm(grid[r][c])
            if part and part not in parts:
                parts.append(part)
        headers.append(" | ".join(parts))
    rows = table_matrix(table)[len(header_trs):]
    return headers, rows


def find_table(html: str, required: Iterable[str], optional: Iterable[str] = ()):
    soup = BeautifulSoup(html, "html.parser")
    choices = []
    required = tuple(required)
    optional = tuple(optional)
    for table in soup.find_all("table"):
        text = norm(table.get_text(" ", strip=True))
        if all(norm(x) in text for x in required):
            headers, rows = expanded_headers(table)
            score = 100 * len(required) + 10 * sum(norm(x) in text for x in optional) + min(len(rows), 20)
            choices.append((score, headers, rows))
    if not choices:
        raise ValueError(f"No APSRTC table found for {list(required)}")
    _, headers, rows = max(choices, key=lambda item: item[0])
    return headers, rows


def col(headers, aliases):
    normalized = [norm(h) for h in headers]
    for alias in aliases:
        a = norm(alias)
        for idx, header in enumerate(normalized):
            if header == a or a in header:
                return idx
    return None


def row_value(headers, row, aliases):
    idx = col(headers, aliases)
    return number(row[idx]) if idx is not None and idx < len(row) else None


def depot_row(headers, rows, display_name, vehicle_depot):
    depot_idx = col(headers, ["DEPOT", "DEPOT NAME", "DISTRICT"])
    wanted = {norm(display_name), norm(vehicle_depot), norm(vehicle_depot.split("/")[-1])}
    for row in rows:
        if depot_idx is not None and depot_idx < len(row) and norm(row[depot_idx]) in wanted:
            return row
        if any(norm(cell) in wanted for cell in row):
            return row
    return None


def source_params(year: int, month: int, region: str, vehicle_depot: str):
    last = monthrange(year, month)[1]
    return {
        "action": "",
        "fdate": f"{year:04d}-{month:02d}-{last:02d}",
        "rreg": region,
        "dept": vehicle_depot,
        "dist": vehicle_depot,
    }


def fetch_hsd_incl(session, display, vehicle, region, y, m):
    html = request_html(session, MED_BASE, "hsdtrend_dpt.php", {
        "action": "", "mth": f"{y}{m:02d}", "rreg": region, "prmonth": f"March-{y}"
    })
    h, rows = find_table(html, ["DEPOT"], ["UP TO THE MONTH", "TGT KMPL", "KMPL"])
    row = depot_row(h, rows, display, vehicle)
    return row_value(h, row, ["UP TO THE MONTH CY", "UPTO THE MONTH CY", "UP TO MONTH CY", "UPTO CY"]) if row else None


def fetch_hsd_excl(session, display, vehicle, region, y, m):
    html = request_html(session, MED_BASE, "achsd1.php", source_params(y, m, region, vehicle))
    try:
        from src.parser.region_parser import parse_region_html
        for rec in parse_region_html(html):
            if norm(rec.get("depot")) in {norm(display), norm(vehicle), norm(vehicle.split("/")[-1])} and norm(rec.get("category")) == "NAC":
                return number(rec.get("upd_kmpl"))
    except Exception:
        pass
    h, rows = find_table(html, ["NAC", "AC"], ["UPD KMPL", "KMPL"])
    di, ci = col(h, ["DEPOT"]), col(h, ["CATEGORY", "TYPE", "BUS TYPE"])
    for row in rows:
        if di is not None and ci is not None and di < len(row) and ci < len(row):
            if norm(row[di]) in {norm(display), norm(vehicle), norm(vehicle.split("/")[-1])} and norm(row[ci]) == "NAC":
                return row_value(h, row, ["UPD KMPL", "UP TO DAY KMPL", "KMPL"])
    return None


def fetch_direct_dimension(session, path: str, label: str, group: str, vehicle, region, y, m):
    html = request_html(session, MEDNEW_BASE, path, source_params(y, m, region, vehicle))
    h, rows = find_table(html, [label], ["FOR THE MONTH", "UP TO THE MONTH"])
    name_idx = col(h, [label])
    upto_idx = col(h, ["UP TO THE MONTH | CY", "UP TO THE MONTH CY", "UPTO THE MONTH CY"])
    if upto_idx is None:
        candidates = [i for i, header in enumerate(h) if "UP TO THE MONTH" in norm(header) and re.search(r"\bCY\b", norm(header))]
        upto_idx = candidates[0] if candidates else None
    output = []
    if name_idx is None or upto_idx is None:
        return output
    for row in rows:
        if name_idx >= len(row) or upto_idx >= len(row):
            continue
        name = row[name_idx].strip()
        if not name or norm(name) in {"TOTAL", "GRAND TOTAL"}:
            continue
        output.append((f"{group}: {name}", number(row[upto_idx]), "0.00"))
    return output


def fetch_lub(session, display, vehicle, region, y, m):
    html = request_html(session, MED_BASE, "lub_rgn_rpt.php", source_params(y, m, region, vehicle))
    h, rows = find_table(html, ["TOTAL LUB KMPL"], ["DEPOT", "DISTRICT", "UPTO THE MONTH CY"])
    row = depot_row(h, rows, display, vehicle)
    return row_value(h, row, ["TOTAL LUB KMPL"]) if row else None


def fetch_breakdown(session, display, vehicle, region, y, m):
    html = request_html(session, MED_BASE, "sysbd_dpt.php", source_params(y, m, region, vehicle))
    h, rows = find_table(html, ["BD RATE"], ["UPTO THE MONTH", "TOTAL BDS"])
    row = depot_row(h, rows, display, vehicle)
    return row_value(h, row, ["UPTO THE MONTH | BD RATE", "UP TO THE MONTH BD RATE", "BD RATE"]) if row else None


def fetch_med(session, display, vehicle, region, y, m):
    params = source_params(y, m, region, vehicle)
    params["dist"] = vehicle
    html = request_html(session, MED_BASE, "medcan_um_dpt.php", params)
    h, rows = find_table(html, ["TOTAL KMS CANC."], ["UP TO THE MONTH", "% FO CANC.", "% CANC."])
    row = depot_row(h, rows, display, vehicle)
    return row_value(h, row, ["UP TO THE MONTH | % FO CANC.", "UPTO THE MONTH % FO CANC.", "% FO CANC.", "% CANC."]) if row else None


def fetch_spring(session, display, vehicle, region, y, m):
    html = request_html(session, MED_BASE, "deptspring.php", source_params(y, m, region, vehicle))
    h, rows = find_table(html, ["SPRING"], ["CONSUMPTION PER LAKH KMS", "UPTO THE MONTH"])
    row = depot_row(h, rows, display, vehicle)
    return row_value(h, row, ["UPTO THE MONTH | SPRING CONSUMPTION PER LAKH KMS", "UP TO THE MONTH SPRING CONSUMPTION PER LAKH KMS", "SPRING CONSUMPTION PER LAKH KMS"]) if row else None


def tyre_total_row(headers, rows, display, vehicle):
    di, si = col(headers, ["DEPOT"]), col(headers, ["TYRE SIZE", "SIZE"])
    wanted = {norm(display), norm(vehicle), norm(vehicle.split("/")[-1])}
    fallback = None
    for row in rows:
        if di is None or di >= len(row) or norm(row[di]) not in wanted:
            continue
        fallback = fallback or row
        if si is not None and si < len(row) and "ALL TYRE" in norm(row[si]):
            return row
    return fallback


def fetch_tyre(session, display, vehicle, region, y, m):
    params = source_params(y, m, region, vehicle)
    params.update({
        "month_year": datetime(y, m, 1).strftime("%b-%Y").upper(),
        "tyre_size": "All Tyre Sizes Total",
        "depot": vehicle.split("/")[-1],
    })
    html = request_html(session, TYRE_BASE, "d_statement_final.php", params)
    h, rows = find_table(html, ["DEPOT", "TYRE SIZE"], ["RT_FACTOR", "NEW MILEAGE", "AVG TOTAL MILEAGE"])
    row = tyre_total_row(h, rows, display, vehicle)
    if not row:
        return {}
    rv = lambda aliases: row_value(h, row, aliases)
    return {
        "AVG TYRE LIFE": (rv(["AVG TOTAL MILEAGE", "AVERAGE TOTAL MILEAGE"]), "lakh"),
        "NEW TYRE LIFE": (rv(["NEW MILEAGE"]), "lakh"),
        "RC TYRE LIFE": (rv(["RC MILEAGE"]), "lakh"),
        "N.T.S RATE": (rv(["NEW %", "NTS", "N.T.S"]), "0.00"),
        "Ist RC S Rate": (rv(["IST RC %", "IST RC SCP %", "1ST RC %"]), "0.00"),
        "TTL SCP Rate": (rv(["TOTAL %", "TTL.SCP %", "TOTAL SCRAP %"]), "0.00"),
        "RT Factor": (rv(["RT_FACTOR", "RT FACTOR"]), "0.00"),
    }


def collect_fy(session, display, vehicle, region, fy):
    y, m = effective_month_for_fy(fy)
    print(f"FY {fy}: cumulative source month {y:04d}-{m:02d}")
    rows = []
    add = lambda name, value, fmt="0.00": rows.append((name, value, fmt))

    add("HSD KMPL INCL AC", fetch_hsd_incl(session, display, vehicle, region, y, m))
    add("HSD KMPL EXCL AC", fetch_hsd_excl(session, display, vehicle, region, y, m))
    rows.extend(fetch_direct_dimension(session, "prodkmpl_um.php", "PRODUCT", "PRODUCT", vehicle, region, y, m))
    rows.extend(fetch_direct_dimension(session, "engkmpl_um.php", "ENGINE TYPE", "ENGINE", vehicle, region, y, m))
    add("TOTAL LUB KMPL", fetch_lub(session, display, vehicle, region, y, m), "0")
    add("B.D RATE", fetch_breakdown(session, display, vehicle, region, y, m))
    add("MED CANCL.", fetch_med(session, display, vehicle, region, y, m))
    add("SPRING CONS", fetch_spring(session, display, vehicle, region, y, m))

    for name, (value, fmt) in fetch_tyre(session, display, vehicle, region, y, m).items():
        if fmt == "lakh" and value is not None:
            value, fmt = value / 100000.0, "0.00"
        add(name, value, fmt)
    return rows


def build_workbook(display: str, financial_years: list[str], fy_rows: dict[str, list[tuple]]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Annual KPI"
    ws.append(["KPI"] + financial_years)

    names, formats = [], {}
    values = {fy: {} for fy in financial_years}
    for fy in financial_years:
        for name, value, fmt in fy_rows[fy]:
            if name not in names:
                names.append(name)
            formats[name] = fmt
            values[fy][name] = value

    for name in names:
        ws.append([name] + [values[fy].get(name) for fy in financial_years])

    ws.freeze_panes = "B2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    for row in ws.iter_rows(min_row=2):
        row[0].font = Font(bold=True)
        row[0].border = BORDER
        for cell in row[1:]:
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
            if cell.value is not None:
                cell.number_format = formats.get(row[0].value, "0.00")

    ws.column_dimensions["A"].width = 38
    for idx in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 16
    ws.insert_rows(1, 2)
    ws["A1"] = f"{display} - ANNUAL KPI DASHBOARD"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(financial_years))
    ws["A1"].alignment = Alignment(horizontal="center")
    return wb


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--depot", required=True)
    parser.add_argument("--fy", action="append", dest="financial_years", help="YYYY-YY; repeat for multiple FYs")
    parser.add_argument("--years", help="Comma-separated FYs, e.g. 2023-24,2024-25,2025-26")
    args = parser.parse_args()

    financial_years = args.financial_years or []
    if args.years:
        financial_years.extend([x.strip() for x in args.years.split(",") if x.strip()])
    if not financial_years:
        financial_years = ["2023-24", "2024-25", "2025-26"]
    financial_years = list(dict.fromkeys(financial_years))
    for fy in financial_years:
        parse_fy(fy)
        effective_month_for_fy(fy)

    vehicle, display, region = depot_info(args.depot)
    session = login()
    fy_rows = {fy: collect_fy(session, display, vehicle, region, fy) for fy in financial_years}

    reports = PROJECT_DIR / "reports"
    reports.mkdir(exist_ok=True)
    year_tag = "_".join(fy.replace("-", "_") for fy in financial_years)
    xlsx_path = reports / f"{display}_ANNUAL_KPI_{year_tag}.xlsx"
    build_workbook(display, financial_years, fy_rows).save(xlsx_path)

    folder_id = os.getenv("KPI_DRIVE_FOLDER_ID", DEFAULT_DRIVE_FOLDER)
    sheet_name = f"{display}_ANNUAL_KPI_{year_tag}"
    uploaded = upload_xlsx_as_google_sheet(xlsx_path, folder_id, sheet_name)
    if uploaded.get("already_existed"):
        print(f"ALREADY_DELIVERED: {uploaded.get('webViewLink')}")
    else:
        print(f"ANNUAL_KPI_SUCCESS: {uploaded.get('webViewLink')}")
    print(f"GOOGLE_SHEET_ID: {uploaded.get('id')}")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as exc:
        print(f"ANNUAL_KPI_FAILURE: {exc}")
        sys.exit(1)
