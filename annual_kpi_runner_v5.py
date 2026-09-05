#!/usr/bin/env python3
import argparse, os, sys
import annual_kpi_runner_v4 as v4

m = v4.m
LAYOUT_VERSION = "5"


def exact_lub_html_v5(s, y, month):
    wanted = f"{m.datetime(y, month, 1).strftime('%B')}_{y}"
    url = f"{m.core.MED_BASE}/lub_rgn_rpt.php"
    rr = s.post(url, data={"dt": wanted}, headers={"Content-Type": "application/x-www-form-urlencoded"}, timeout=45)
    rr.raise_for_status()
    text = rr.text
    if "<table" not in text.lower() or "TOTAL LUB KMPL" not in m.n(text):
        raise RuntimeError(f"LUB POST dt returned no Total Lub table: {wanted}")
    print(f"LUB POST accepted {wanted} using field: dt")
    return text


def format_sheet_v5(spreadsheet_id, mat, fys):
    v4.clean_format_sheet_17(spreadsheet_id, mat, fys)
    sid = m.sheet_id(spreadsheet_id, m.SHEET_TITLE)
    svc = m.sheets_service()
    req = []
    for r in range(1, len(mat)):
        block_start = r - ((r - 1) % 3)
        kpi = str(mat[block_start][1])
        pattern = "0" if kpi == "TOTAL LUB KMPL" else "0.00"
        for c0, c1 in ((3, 15), (16, 17)):
            req.append({"repeatCell": {"range": {"sheetId": sid, "startRowIndex": r, "endRowIndex": r + 1, "startColumnIndex": c0, "endColumnIndex": c1}, "cell": {"userEnteredFormat": {"numberFormat": {"type": "NUMBER", "pattern": pattern}}}, "fields": "userEnteredFormat.numberFormat"}})
    if req:
        svc.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body={"requests": req}).execute()


def make_xlsx_v5(display, mat, fys):
    path = v4.make_xlsx_17(display, mat, fys)
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb[m.SHEET_TITLE]
    for r in range(2, ws.max_row + 1):
        block_start = r - ((r - 2) % 3)
        kpi = str(ws.cell(block_start, 2).value or "")
        pattern = "0" if kpi == "TOTAL LUB KMPL" else "0.00"
        for c in list(range(4, 16)) + [17]:
            ws.cell(r, c).number_format = pattern
    wb.save(path)
    return path


def meta_version(spreadsheet_id):
    try:
        vals = m.read_values(spreadsheet_id, f"'{m.META_TITLE}'!A:B")
        d = {str(r[0]): str(r[1]) for r in vals[1:] if len(r) >= 2}
        return d.get("LAYOUT_VERSION", "")
    except Exception:
        return ""


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--depot", required=True)
    p.add_argument("--selected-month", required=True)
    p.add_argument("--years", default="")
    a = p.parse_args()
    sy, sm = map(int, a.selected_month.split("-"))
    m.datetime(sy, sm, 1)
    fys = m.fy_triplet(sy, sm)
    vehicle, display, region = m.core.depot_info(a.depot)
    folder = os.getenv("KPI_DRIVE_FOLDER_ID", m.core.DEFAULT_DRIVE_FOLDER)
    name = f"{display}_ANNUAL_KPI_DASHBOARD"
    existing = m.find_file(folder, name)
    repair = (not existing) or meta_version(existing["id"]) != LAYOUT_VERSION

    # IMPORTANT: on a version repair rebuild from APSRTC source instead of trusting a shifted/corrupted sheet.
    if repair:
        st = m.new_store(fys)
        m.seed(st)
    else:
        st = v4.load_existing_shifted(m.read_values(existing["id"], f"'{m.SHEET_TITLE}'!A:Q"), fys)
        m.seed(st)

    s = m.login()
    if repair:
        print(f"FULL REPAIR/BUILD {display}: {', '.join(fys)}")
        for fy in fys:
            months = m.fy_months(fy, sy, sm)
            for i, (y, mon) in enumerate(months):
                need = i == len(months) - 1
                print(f"FETCH {fy} {y:04d}-{mon:02d}{' + Upto' if need else ''}")
                m.apply(st, fy, y, mon, m.fetch_month(s, display, vehicle, region, y, mon, need), need, True)
    else:
        fy = m.selected_fy(sy, sm)
        print(f"INCREMENTAL {display} {a.selected_month}")
        m.apply(st, fy, sy, sm, m.fetch_month(s, display, vehicle, region, sy, sm, True), True, False)

    mat = v4.matrix_with_serial(st)
    xlsx = make_xlsx_v5(display, mat, fys)
    if not existing:
        up = m.upload_xlsx_as_google_sheet(xlsx, folder, name)
        sid = up["id"]
        link = up.get("webViewLink", "")
    else:
        sid = existing["id"]
        link = existing.get("webViewLink", "")

    format_sheet_v5(sid, mat, fys)
    m.ensure_hidden_sheet(sid, m.META_TITLE)
    m.write_values(sid, f"'{m.META_TITLE}'!A1", [["KEY","VALUE"],["DEPOT",display],["FYS",','.join(fys)],["LAST_SELECTED_MONTH",a.selected_month],["LAYOUT_VERSION",LAYOUT_VERSION]])
    print(f"ANNUAL_KPI_DASHBOARD_SUCCESS: {link}")
    print(f"GOOGLE_SHEET_ID: {sid}")
    return 0


m.lub_html = exact_lub_html_v5
m.format_sheet = format_sheet_v5
m.make_xlsx = make_xlsx_v5

if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as exc:
        print(f"ANNUAL_KPI_FAILURE: {exc}")
        raise
